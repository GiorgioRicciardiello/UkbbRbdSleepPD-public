"""
Generate publication-ready formatted Cox model tables.

Reads raw Excel outputs from a results/cox_prodromal_abk_* directory and
writes a cleaned, formatted workbook to report_formal/ in the same directory.

Tables generated (Supplementary Table numbering matches manuscript)
-------------------------------------------------------------------
  Supp T2  — RBD-only HRs (Model A, percentile_3g, all outcomes)
  Supp T3  — RBD threshold stability (PD only)
  Supp T4  — Prodromal-only HRs (Model B, PD only, FDR-corrected)
  Supp T5  — Additive model (Model C) — PD only, percentile_3g
  Supp T6  — Incremental C-index by prodromal marker (PD only)
  Supp T7  — Interaction terms (Model D) — PD only, percentile_3g
  Supp T9  — Competing-risk Cox (Model E) — PD only
  Supp T10 — Lag sensitivity (2-year exclusion) — PD only
  Supp T11 — HES-active subcohort sensitivity — PD only
  Supp T12a — RBD spline LRT (non-linearity) — PD only
  Supp T12b — Prodromal spline LRT (linearity) — PD only
  Extra    — RBD continuous HR per 1-SD (all outcomes)
  Extra    — CIF vs KM comparison — PD only
  Extra    — PH diagnostics — PD only, primary model
  Extra    — KM log-rank summary — PD only, percentile_3g
  Extra    — Absolute risks by group — PD only
  Summary  — Combined key results, one row per (outcome x model x exposure)

Scope rules applied
-------------------
  PD_ONLY tables  : filtered to outcome_1a_pd_only
  3G_ONLY tables  : filtered to method == PRIMARY_METHOD ("percentile_3g")
  ALL_OUTCOMES    : rbd_only_cox and rbd_continuous (needed for specificity claim)

Usage
-----
  python generate_formal_tables.py
  # or override path at top of file
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
sys.path.insert(0, str(Path(__file__).resolve().parent))
from config.config import outcomes as _cfg_outcomes, outcomes_formal_names as _cfg_formal  # noqa: E402

# ── Configuration — edit TARGET_DIR to point at a different run ────────────
TARGET_DIR: Path = Path(
    "results/cox_prodromal_abk_03_19_2026_14_17_29"
)
PRIMARY_METHOD: str = "percentile_3g"   # "percentile_3g" | "percentile_2g"
PRIMARY_OUTCOME: str = "outcome_1a_pd_only"
OUTPUT_SUBDIR: str  = "report_formal"

# ── Label mappings (sourced from config.config — single source of truth) ───
OUTCOME_LABELS: Dict[str, str] = _cfg_formal
OUTCOME_ORDER: List[str] = list(_cfg_outcomes)

COVARIATE_LABELS: Dict[str, str] = {
    "cov_age_recruitment_21022":       "Age at recruitment",
    "cov_sex_31":                      "Sex (male)",
    "bmi_imp_23104_i0":                "BMI",
    "bmi_imp_23104_bl":                "BMI",
    "bmi_21001_bl":                    "BMI",
    "cov_smoking":                     "Smoking",
    "cov_alcohol":                     "Alcohol use",
    # RBD groups
    "rbd_High (99,100%)":              "RBD — High (99–100th pctile)",
    "rbd_Intermediate (90,99%)":       "RBD — Intermediate (90–99th pctile)",
    "rbd_High (90,100%)":              "RBD — High (90–100th pctile)",
    # Prodromal group labels
    "prod_High":                       "Prodromal — High",
    "prod_Medium":                     "Prodromal — Medium",
    "prod_Yes":                        "Prodromal — Yes",
    # Prodromal vars (legacy _i0 keys retained for old result files)
    "cov_fluid_intelligence_20016_i0":     "Fluid Intelligence",
    "cov_react_time_mean_20023_i0":        "Reaction Time (ms)",
    "cov_fi_questions_attempted_20128_i0": "FI Questions Attempted",
    "cov_numeric_memory_max_20240_i0":     "Numeric Memory",
    "trail_making_errors_trail1_i2":       "Trail Making",
    "cov_pairs_status_20244_i0":           "Pairs Matching",
    "prodromal_constipation":         "Constipation",
    "prodromal_depression":           "Depression",
    "prodromal_anxiety":              "Anxiety",
    "prodromal_orthostatic":          "Orthostatic Hypotension",
    "prodromal_erectile_dysfunction": "Erectile Dysfunction",
    "prodromal_dream_enactment":      "Dream Enactment",
    "prodromal_anosmia":              "Anosmia",
    "prodromal_hyposmia":             "Hyposmia",
    # Prodromal vars — current _bl/_fu names (post temporal-window rename).
    # trail_making_errors_trail1_i2 keeps _i2 (derived from p6348, not renamed).
    "cog_fluid_intelligence_bl":           "Fluid Intelligence",
    "cog_react_time_bl":                   "Reaction Time (ms)",
    "cov_fi_questions_attempted_20128_bl": "FI Questions Attempted",
    "cog_numeric_memory_bl":               "Numeric Memory",
    "cog_pairs_matching_bl":               "Pairs Matching",
    "cog_tmt_ratio_log_bl":                "TMT-B/A Ratio (log)",
    "prodromal_constipation_bl":         "Constipation",
    "prodromal_depression_bl":           "Depression",
    "prodromal_anxiety_bl":              "Anxiety",
    "prodromal_orthostatic_bl":          "Orthostatic Hypotension",
    "prodromal_erectile_dysfunction_bl": "Erectile Dysfunction",
    "prodromal_dream_enactment_bl":      "Dream Enactment",
    "prodromal_anosmia_bl":              "Anosmia",
    "prodromal_hyposmia_bl":             "Hyposmia",
    # Incident post-baseline prodromal markers
    "prodromal_constipation_post":         "Constipation (incident)",
    "prodromal_depression_post":           "Depression (incident)",
    "prodromal_anxiety_post":              "Anxiety (incident)",
    "prodromal_orthostatic_post":          "Orthostatic Hypotension (incident)",
    "prodromal_erectile_dysfunction_post": "Erectile Dysfunction (incident)",
    "prodromal_burden_post":               "Prodromal burden (incident, post-baseline)",
}

MODEL_LABELS: Dict[str, str] = {
    "M0_rbd_only":     "Model A — RBD only",
    "M1_prodromal":    "Model B — Prodromal only",
    "M2_additive":     "Model C — Additive (RBD + Prodromal)",
    "M3_interaction":  "Model D — Interaction",
    "M4_competing":    "Model E — Competing risks",
}

MODEL_ORDER: List[str] = [
    "M0_rbd_only", "M1_prodromal", "M2_additive",
    "M3_interaction", "M4_competing",
]

# ── Rounding / formatting helpers ──────────────────────────────────────────

def _fmt_p(p: float) -> str:
    """Format p-value to 3 decimals; use '< 0.001' for very small values."""
    if pd.isna(p):
        return "—"
    if p < 0.001:
        return "< 0.001"
    return f"{p:.3f}"


def _fmt_hr(hr: float, lo: float, hi: float) -> str:
    """Format HR with 95% CI as 'HR [lo, hi]'."""
    if any(pd.isna(x) for x in [hr, lo, hi]):
        return "—"
    return f"{hr:.2f} [{lo:.2f}, {hi:.2f}]"


def _fmt_ci(lo: float, hi: float) -> str:
    if any(pd.isna(x) for x in [lo, hi]):
        return "—"
    return f"[{lo:.2f}, {hi:.2f}]"


def _fmt_pct(x: float) -> str:
    if pd.isna(x):
        return "—"
    return f"{x:.2f}%"


# ── Label application helpers ──────────────────────────────────────────────

def _label_outcome(s: pd.Series) -> pd.Series:
    return s.map(lambda x: OUTCOME_LABELS.get(x, x))


def _label_covariate(s: pd.Series) -> pd.Series:
    return s.map(lambda x: COVARIATE_LABELS.get(str(x), str(x)))


def _order_outcomes(df: pd.DataFrame, col: str = "outcome") -> pd.DataFrame:
    """Sort DataFrame rows by the canonical outcome order."""
    order = {o: i for i, o in enumerate(OUTCOME_ORDER)}
    df = df.copy()
    df["_order"] = df[col].map(order).fillna(len(OUTCOME_ORDER))
    df = df.sort_values("_order").drop(columns=["_order"])
    return df


# ── Scope filters ──────────────────────────────────────────────────────────

def _filter_pd_3g(df: pd.DataFrame) -> pd.DataFrame:
    """Keep PD-only outcome and primary method rows."""
    if "outcome" in df.columns:
        df = df[df["outcome"] == PRIMARY_OUTCOME]
    if "method" in df.columns:
        df = df[df["method"] == PRIMARY_METHOD]
    return df.copy()


def _filter_pd(df: pd.DataFrame) -> pd.DataFrame:
    """Keep PD-only outcome rows (any method)."""
    if "outcome" in df.columns:
        df = df[df["outcome"] == PRIMARY_OUTCOME]
    return df.copy()


# ── Row-type filters ───────────────────────────────────────────────────────

def _is_rbd_row(cov: str) -> bool:
    return cov.startswith("rbd_")


def _is_prodromal_row(cov: str) -> bool:
    """Prodromal exposure rows (not baseline covariates, not RBD)."""
    skip = {
        "cov_age_recruitment_21022", "cov_sex_31",
        "bmi_imp_23104_i0", "bmi_imp_23104_bl", "cov_bmi",
        "cov_smoking", "cov_alcohol",
    }
    return cov not in skip and not cov.startswith("rbd_")


def _is_interaction_row(cov: str) -> bool:
    return "__x__" in cov


# ── Table builders ─────────────────────────────────────────────────────────

def build_supp_t2_rbd_only(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 2: RBD-only Cox HRs (Model A, percentile_3g) — all outcomes.

    Scope: all outcomes, PRIMARY_METHOD only, RBD exposure rows.
    Shows the outcome-specificity gradient (PD >> AD/dementia).
    """
    df = df_raw.copy()
    if "method" in df.columns:
        df = df[df["method"] == PRIMARY_METHOD]
    df = _order_outcomes(df)

    df["Outcome"]     = _label_outcome(df["outcome"])
    df["Covariate"]   = _label_covariate(df["covariate"])
    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    # Keep RBD exposure rows only for the main display
    rbd_mask = df["covariate"].apply(_is_rbd_row)
    df = df[rbd_mask]

    return df[[
        "Outcome", "Covariate", "N", "events",
        "HR [95% CI]", "p-value",
    ]].rename(columns={"events": "Events"})


def build_supp_t3_threshold_stability(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 3: RBD threshold stability across percentile cutoffs (PD only).

    Scope: PD-only outcome, top-5/10/15th percentile alternatives.
    """
    df = _filter_pd(df_raw.copy())

    df["Percentile (%)"]  = df["percentile"].map(lambda x: f"{x:.0f}%")
    df["HR [95% CI]"]     = df.apply(
        lambda r: _fmt_hr(r["hr"], r["lci"], r["uci"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    cols = ["Percentile (%)"]
    if "threshold_value" in df.columns:
        df["Score threshold"] = df["threshold_value"].map(
            lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
        )
        cols.append("Score threshold")
    cols += ["HR [95% CI]", "p-value"]
    if "n_high" in df.columns:
        df["N (high)"] = df["n_high"].astype(int)
        cols.append("N (high)")
    if "events" in df.columns:
        df["Events"] = df["events"].astype(int)
        cols.append("Events")

    df = df.sort_values("percentile") if "percentile" in df.columns else df
    return df[[c for c in cols if c in df.columns]]


def build_supp_t4_prodromal_only(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 4: Prodromal-only HRs (Model B) — PD only, exposure rows, FDR-corrected.

    Scope: PD-only outcome. Prodromal exposure rows only (no baseline covariates).
    """
    df = _filter_pd(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["Covariate"] = _label_covariate(df["covariate"])

    # Keep exposure rows
    exp_mask = df["covariate"].apply(_is_prodromal_row)
    df = df[exp_mask]

    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)
    df["p_FDR"]   = df["p_fdr"].apply(_fmt_p) if "p_fdr" in df.columns else "—"
    df["N"]       = df["N"].astype(int)
    df["Events"]  = df["events"].astype(int)

    df = df.sort_values(["prodromal_var", "HR"], ascending=[True, False])

    return df[[
        "Prodromal marker", "Covariate", "N", "Events",
        "HR [95% CI]", "p-value", "p_FDR",
    ]]


def build_supp_t5_additive(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 5: Additive model (Model C) — PD only, percentile_3g, exposure rows.

    Scope: PD-only, PRIMARY_METHOD, M2_additive.
    Shows RBD HR stability after conditioning on each prodromal marker.
    """
    df = df_raw.copy()
    df = df[df["model"] == "M2_additive"]
    df = _filter_pd_3g(df)

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["Covariate"]   = _label_covariate(df["covariate"])
    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    # Keep RBD and prodromal exposure rows; drop baseline covariates
    exp_mask = df["covariate"].apply(
        lambda c: _is_rbd_row(c) or _is_prodromal_row(c)
    )
    df = df[exp_mask]

    df = df.sort_values(["prodromal_var", "covariate"])

    return df[[
        "Prodromal marker", "Covariate", "N", "events",
        "HR [95% CI]", "p-value",
    ]].rename(columns={"events": "Events"})


def build_supp_t6_c_index(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 6: Incremental C-index by prodromal marker — PD only.

    Scope: PD-only outcome.
    """
    df = _filter_pd(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["C-index (full)"] = df["c_index_full"].map(
        lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
    )
    df["C-index (null)"] = df["c_index_null"].map(
        lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
    )
    df["Delta C-index"] = df["c_index_incremental"].map(
        lambda x: f"{x:+.4f}" if not pd.isna(x) else "—"
    )

    return df[[
        "Prodromal marker",
        "C-index (full)", "C-index (null)", "Delta C-index",
    ]]


def build_supp_t7_interaction(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 7: Interaction model (Model D) — PD only, percentile_3g, interaction rows.

    Scope: PD-only, PRIMARY_METHOD, M3_interaction, interaction term rows only.
    """
    df = df_raw.copy()
    df = df[df["model"] == "M3_interaction"]
    df = _filter_pd_3g(df)

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["Covariate"] = df["covariate"]  # keep raw; interaction rows contain __x__

    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    # Interaction rows only
    ix_mask = df["covariate"].apply(_is_interaction_row)
    df = df[ix_mask]

    df = df.sort_values(["prodromal_var", "HR"], ascending=[True, False])

    return df[[
        "Prodromal marker", "Covariate", "N", "events",
        "HR [95% CI]", "p-value",
    ]].rename(columns={"events": "Events"})


def build_supp_t9_competing_risk(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 9: Competing-risk Cox (Model E) — PD only.

    Scope: PD-only outcome. Cause-specific HR for RBD exposure.
    """
    df = _filter_pd(df_raw.copy())

    df["Covariate"]   = _label_covariate(df["covariate"])
    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    cols = ["Covariate", "HR [95% CI]", "p-value"]
    if "N" in df.columns:
        cols = ["Covariate", "N"] + cols[1:]

    return df[[c for c in cols if c in df.columns]]


def build_supp_t10_lag_sensitivity(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 10: Lag sensitivity (2-year exclusion) — PD only, exposure rows.

    Columns: primary HR vs lag-2yr HR for each prodromal exposure.
    """
    df = _filter_pd(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["Covariate"] = _label_covariate(df["covariate"])

    # Keep prodromal exposure rows only
    exp_mask = df["covariate"].apply(_is_prodromal_row)
    df = df[exp_mask]

    df["HR (primary)"] = df["HR_primary"].map(
        lambda x: f"{x:.2f}" if not pd.isna(x) else "—"
    )
    df["HR [95% CI] (lag 2y)"] = df.apply(
        lambda r: _fmt_hr(r["HR_lag2y"], r["HR_lag2y_lower"], r["HR_lag2y_upper"]), axis=1
    )
    df["p-value (lag 2y)"] = df["p_lag2y"].apply(_fmt_p)

    cols = ["Prodromal marker", "Covariate", "HR (primary)",
            "HR [95% CI] (lag 2y)", "p-value (lag 2y)"]
    if "N_lag" in df.columns:
        df["N (lag)"] = df["N_lag"].astype(int)
        cols.append("N (lag)")
    if "events_lag" in df.columns:
        df["Events (lag)"] = df["events_lag"].astype(int)
        cols.append("Events (lag)")

    df = df.sort_values(["prodromal_var", "covariate"])
    return df[[c for c in cols if c in df.columns]]


def build_supp_t11_hes_sensitivity(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 11: HES-active subcohort sensitivity — PD only, exposure rows.

    Restricted to participants with HES gap ≤ 4 years pre-baseline.
    """
    df = _filter_pd(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )
    df["Covariate"] = _label_covariate(df["covariate"])

    exp_mask = df["covariate"].apply(_is_prodromal_row)
    df = df[exp_mask]

    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["HR"], r["HR_lower"], r["HR_upper"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)

    cols = ["Prodromal marker", "Covariate", "HR [95% CI]", "p-value"]
    if "N_sensitivity" in df.columns:
        df["N (HES-active)"] = df["N_sensitivity"].astype(int)
        cols = ["Prodromal marker", "Covariate", "N (HES-active)"] + cols[2:]
    if "events_sensitivity" in df.columns:
        df["Events (HES-active)"] = df["events_sensitivity"].astype(int)
        cols.append("Events (HES-active)")
    if "N_excluded_gap" in df.columns:
        df["N excluded (gap > 4y)"] = df["N_excluded_gap"].astype(int)
        cols.append("N excluded (gap > 4y)")

    df = df.sort_values(["prodromal_var", "covariate"])
    return df[[c for c in cols if c in df.columns]]


def build_supp_t12a_rbd_spline(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 12a: RBD spline LRT — non-linearity test (PD only).

    Tests spline vs linear RBD for the primary outcome.
    Columns: LRT statistic, p-value, C-index (spline vs linear).
    """
    df = _filter_pd(df_raw.copy())

    df["Outcome"] = _label_outcome(df["outcome"])

    cols = ["Outcome"]
    if "N" in df.columns:
        df["N"] = df["N"].astype(int)
        cols.append("N")
    if "events" in df.columns:
        df["Events"] = df["events"].astype(int)
        cols.append("Events")
    if "lr_stat" in df.columns:
        df["LRT statistic"] = df["lr_stat"].map(
            lambda x: f"{x:.2f}" if not pd.isna(x) else "—"
        )
        cols.append("LRT statistic")
    if "lr_p" in df.columns:
        df["LRT p-value"] = df["lr_p"].apply(_fmt_p)
        cols.append("LRT p-value")
    if "c_index_spline" in df.columns:
        df["C-index (spline)"] = df["c_index_spline"].map(
            lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
        )
        cols.append("C-index (spline)")
    if "c_index_linear" in df.columns:
        df["C-index (linear)"] = df["c_index_linear"].map(
            lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
        )
        cols.append("C-index (linear)")

    return df[[c for c in cols if c in df.columns]]


def build_supp_t12b_prodromal_spline(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Supp Table 12b: Prodromal spline LRT — linearity test (PD only).

    For each continuous prodromal marker, tests spline vs linear fit.
    Non-significant p-values support the linear specification.
    """
    df = _filter_pd(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )

    cols = ["Prodromal marker"]
    if "N" in df.columns:
        df["N"] = df["N"].astype(int)
        cols.append("N")
    if "events" in df.columns:
        df["Events"] = df["events"].astype(int)
        cols.append("Events")
    if "lr_stat" in df.columns:
        df["LRT statistic"] = df["lr_stat"].map(
            lambda x: f"{x:.2f}" if not pd.isna(x) else "—"
        )
        cols.append("LRT statistic")
    if "lr_p" in df.columns:
        df["LRT p (non-linearity)"] = df["lr_p"].apply(_fmt_p)
        cols.append("LRT p (non-linearity)")
    if "c_index_spline" in df.columns:
        df["C-index (spline)"] = df["c_index_spline"].map(
            lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
        )
        cols.append("C-index (spline)")

    df = df.sort_values("prodromal_var") if "prodromal_var" in df.columns else df
    return df[[c for c in cols if c in df.columns]]


def build_extra_rbd_continuous(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extra table: RBD continuous HR per 1-SD — all outcomes.

    Scope: all outcomes (needed for the forest plot specificity claim in Fig 1b).
    """
    df = _order_outcomes(df_raw.copy())

    df["Outcome"]     = _label_outcome(df["outcome"])
    df["HR [95% CI]"] = df.apply(
        lambda r: _fmt_hr(r["hr_per_sd"], r["hr_lci"], r["hr_uci"]), axis=1
    )
    df["p-value"] = df["p"].apply(_fmt_p)
    df["C-index"] = df["c_index"].map(
        lambda x: f"{x:.3f}" if not pd.isna(x) else "—"
    )

    return df[[
        "Outcome", "N", "events",
        "HR [95% CI]", "p-value", "C-index",
    ]].rename(columns={"events": "Events"})


def build_extra_cif_vs_km(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extra table: CIF vs 1-KM comparison at 5- and 10-year horizons — PD only.

    Demonstrates negligible competing-risk bias in the primary KM estimates.
    """
    df = _filter_pd(df_raw.copy())

    df["Group"]     = df["group"]
    df["Timepoint"] = df["timepoint"].map(
        lambda x: f"{x}-year" if not pd.isna(x) else "—"
    )

    cols = ["Group", "Timepoint"]
    for col, label in [
        ("CIF_AJ_pct", "CIF (Aalen-Johansen, %)"),
        ("CIF_KM_pct", "1−KM (%)"),
        ("difference_pct", "Difference (KM − AJ, %)"),
    ]:
        if col in df.columns:
            df[label] = df[col].map(
                lambda x: _fmt_pct(x) if not pd.isna(x) else "—"
            )
            cols.append(label)

    df = df.sort_values(["group", "timepoint"]) if "timepoint" in df.columns else df
    return df[[c for c in cols if c in df.columns]]


def build_extra_ph_diagnostics(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extra table: PH diagnostics (Schoenfeld test) — PD only, primary model (M0).

    Shows which covariates violated PH and whether violations were consequential.
    """
    df = _filter_pd(df_raw.copy())

    # Keep primary RBD-only model results (M0 or M0_rbd_only)
    if "model" in df.columns:
        df = df[df["model"].isin(["M0", "M0_rbd_only", "rbd_only"])]

    df["Covariate"] = _label_covariate(df["covariate"])
    if "prodromal_label" in df.columns:
        df["Prodromal marker"] = df["prodromal_label"].fillna("—")
    else:
        df["Prodromal marker"] = "—"

    cols = ["Prodromal marker", "Covariate"]
    if "ph_stat" in df.columns:
        df["Schoenfeld stat"] = df["ph_stat"].map(
            lambda x: f"{x:.2f}" if not pd.isna(x) else "—"
        )
        cols.append("Schoenfeld stat")
    if "ph_p" in df.columns:
        df["p-value (PH test)"] = df["ph_p"].apply(_fmt_p)
        cols.append("p-value (PH test)")
    if "ph_violation" in df.columns:
        df["PH violation"] = df["ph_violation"].map(
            lambda x: "Yes" if x else "No"
        )
        cols.append("PH violation")

    df = df.sort_values("ph_p") if "ph_p" in df.columns else df
    return df[[c for c in cols if c in df.columns]]


def build_extra_km_logrank(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extra table: KM log-rank summary — PD only, percentile_3g.

    Log-rank p-values for RBD groups and each prodromal marker.
    """
    df = _filter_pd_3g(df_raw.copy())

    df["Prodromal marker"] = df["prodromal_label"].fillna(
        df["prodromal_var"].map(COVARIATE_LABELS)
    )

    cols = ["Prodromal marker"]
    if "N_cc" in df.columns:
        df["N (complete case)"] = df["N_cc"].astype(int)
        cols.append("N (complete case)")
    if "events_cc" in df.columns:
        df["Events (complete case)"] = df["events_cc"].astype(int)
        cols.append("Events (complete case)")
    if "logrank_rbd_p" in df.columns:
        df["Log-rank p (RBD)"] = df["logrank_rbd_p"].apply(_fmt_p)
        cols.append("Log-rank p (RBD)")
    if "logrank_prod_p" in df.columns:
        df["Log-rank p (prodromal)"] = df["logrank_prod_p"].apply(_fmt_p)
        cols.append("Log-rank p (prodromal)")

    df = df.sort_values("prodromal_var") if "prodromal_var" in df.columns else df
    return df[[c for c in cols if c in df.columns]]


def build_extra_absolute_risks(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extra table: Absolute cumulative incidence by RBD group — PD only.

    Kaplan-Meier derived cumulative incidence at 5 and 10 years by group.
    """
    df = _filter_pd(df_raw.copy())

    # Keep rows without prodromal stratification (RBD-group marginal estimates)
    if "prodromal_var" in df.columns:
        marginal_mask = df["prodromal_var"].isna() | (df["prodromal_var"] == "")
        df_marginal = df[marginal_mask].copy()
        if df_marginal.empty:
            df_marginal = df.copy()
    else:
        df_marginal = df.copy()

    df_marginal["Group"]     = df_marginal["group"]
    df_marginal["Timepoint"] = df_marginal["timepoint_years"].map(
        lambda x: f"{x:.0f}-year" if not pd.isna(x) else "—"
    )

    cols = ["Group", "Timepoint"]
    if "cum_inc_pct" in df_marginal.columns:
        df_marginal["Cum. incidence (%)"] = df_marginal["cum_inc_pct"].map(
            lambda x: _fmt_pct(x) if not pd.isna(x) else "—"
        )
        cols.append("Cum. incidence (%)")
    if "ci_lower_pct" in df_marginal.columns and "ci_upper_pct" in df_marginal.columns:
        df_marginal["95% CI (%)"] = df_marginal.apply(
            lambda r: _fmt_ci(r["ci_lower_pct"], r["ci_upper_pct"]), axis=1
        )
        cols.append("95% CI (%)")
    if "n" in df_marginal.columns:
        df_marginal["N"] = df_marginal["n"].astype(int)
        cols.append("N")
    if "events" in df_marginal.columns:
        df_marginal["Events"] = df_marginal["events"].astype(int)
        cols.append("Events")

    df_marginal = (
        df_marginal.sort_values(["group", "timepoint_years"])
        if "timepoint_years" in df_marginal.columns else df_marginal
    )
    return df_marginal[[c for c in cols if c in df_marginal.columns]]


def build_summary(
    df_rbd_cont: pd.DataFrame,
    df_supp_t2: pd.DataFrame,
    df_supp_t4: pd.DataFrame,
) -> pd.DataFrame:
    """
    Combined summary: key HR rows from continuous RBD, RBD group (PD), prodromal-only (PD).
    """
    rows = []

    # RBD group HRs (all outcomes, 3g)
    if df_supp_t2 is not None and not df_supp_t2.empty:
        pd_only = df_supp_t2[df_supp_t2.get("Outcome", pd.Series(dtype=str)) == "PD only"] \
            if "Outcome" in df_supp_t2.columns else df_supp_t2
        for _, r in pd_only.iterrows():
            rows.append({
                "Model":       "A — RBD groups",
                "Outcome":     r.get("Outcome", "PD only"),
                "Variable":    r.get("Covariate", ""),
                "N":           r.get("N", ""),
                "Events":      r.get("Events", ""),
                "HR [95% CI]": r.get("HR [95% CI]", ""),
                "p-value":     r.get("p-value", ""),
            })

    # Continuous RBD (all outcomes)
    if df_rbd_cont is not None and not df_rbd_cont.empty:
        for _, r in df_rbd_cont.iterrows():
            rows.append({
                "Model":       "A — RBD continuous (per 1-SD)",
                "Outcome":     r.get("Outcome", ""),
                "Variable":    "RBD probability score",
                "N":           r.get("N", ""),
                "Events":      r.get("Events", ""),
                "HR [95% CI]": r.get("HR [95% CI]", ""),
                "p-value":     r.get("p-value", ""),
            })

    # Prodromal-only (PD outcome)
    if df_supp_t4 is not None and not df_supp_t4.empty:
        for _, r in df_supp_t4.iterrows():
            rows.append({
                "Model":       "B — Prodromal only",
                "Outcome":     "PD only",
                "Variable":    f"{r.get('Prodromal marker', '')} [{r.get('Covariate', '')}]",
                "N":           r.get("N", ""),
                "Events":      r.get("Events", ""),
                "HR [95% CI]": r.get("HR [95% CI]", ""),
                "p-value":     r.get("p-value", ""),
            })

    return pd.DataFrame(rows)


# ── Excel formatting ───────────────────────────────────────────────────────

def _write_sheet(
    writer: pd.ExcelWriter,
    df: Optional[pd.DataFrame],
    sheet_name: str,
    title: str = "",
) -> None:
    """Write DataFrame to sheet with basic formatting."""
    if df is None or df.empty:
        warnings.warn(f"Skipping empty table: {sheet_name}")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1 if title else 0)

    ws = writer.sheets[sheet_name]

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    header_row = 2 if title else 1
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Alternate row shading
    light_fill = PatternFill(start_color="EBF0F7", end_color="EBF0F7", fill_type="solid")
    data_start = header_row + 1
    for row_idx, _ in enumerate(df.itertuples(), start=data_start):
        if (row_idx - data_start) % 2 == 0:
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    # Auto-width columns
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            len(str(col_name)),
            *[len(str(v)) for v in df[col_name].values],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)


# ── Main ───────────────────────────────────────────────────────────────────

def main(results_dir: Path = TARGET_DIR) -> None:
    """Build and save all formal tables."""
    results_dir = Path(results_dir)
    if not results_dir.exists():
        print(f"ERROR: directory not found: {results_dir}", file=sys.stderr)
        sys.exit(1)

    out_dir = results_dir / OUTPUT_SUBDIR
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Input  : {results_dir}")
    print(f"Output : {out_dir}")

    # ── Load raw tables ────────────────────────────────────────────────
    def _load(name: str) -> Optional[pd.DataFrame]:
        path = results_dir / f"{name}.xlsx"
        if not path.exists():
            warnings.warn(f"Missing file: {path}. Table will be skipped.")
            return None
        try:
            df = pd.read_excel(path)
            print(f"  Loaded {name}.xlsx  ({len(df)} rows)")
            return df
        except Exception as exc:
            warnings.warn(f"Failed to read {path}: {exc}")
            return None

    # Core tables
    df_rbd_only     = _load("rbd_only_cox")
    df_rbd_cont     = _load("rbd_continuous")
    df_prod_only    = _load("baseline_cox_HRs")
    df_additive     = _load("additive_cox")
    df_interact     = _load("interaction_cox")
    df_competing    = _load("competing_risk_cox")
    df_cindex       = _load("c_index")
    df_stability    = _load("rbd_threshold_stability")
    # Sensitivity / extra tables
    df_lag          = _load("lag_sensitivity")
    df_hes          = _load("sensitivity_hes_active")
    df_rbd_spline   = _load("rbd_spline")
    df_prod_spline  = _load("spline_cox")
    df_cif_vs_km    = _load("competing_risk_cif_vs_km")
    df_ph           = _load("ph_diagnostics")
    df_km_lr        = _load("km_logrank_summary")
    df_abs_risks    = _load("absolute_risks")

    # ── Build tables ───────────────────────────────────────────────────
    t_s2  = build_supp_t2_rbd_only(df_rbd_only)        if df_rbd_only    is not None else None
    t_s3  = build_supp_t3_threshold_stability(df_stability) if df_stability is not None else None
    t_s4  = build_supp_t4_prodromal_only(df_prod_only)  if df_prod_only   is not None else None
    t_s5  = build_supp_t5_additive(df_additive)         if df_additive    is not None else None
    t_s6  = build_supp_t6_c_index(df_cindex)            if df_cindex      is not None else None
    t_s7  = build_supp_t7_interaction(df_interact)      if df_interact    is not None else None
    t_s9  = build_supp_t9_competing_risk(df_competing)  if df_competing   is not None else None
    t_s10 = build_supp_t10_lag_sensitivity(df_lag)      if df_lag         is not None else None
    t_s11 = build_supp_t11_hes_sensitivity(df_hes)      if df_hes         is not None else None
    t_s12a = build_supp_t12a_rbd_spline(df_rbd_spline)  if df_rbd_spline  is not None else None
    t_s12b = build_supp_t12b_prodromal_spline(df_prod_spline) if df_prod_spline is not None else None
    t_cont = build_extra_rbd_continuous(df_rbd_cont)    if df_rbd_cont    is not None else None
    t_cif  = build_extra_cif_vs_km(df_cif_vs_km)        if df_cif_vs_km   is not None else None
    t_ph   = build_extra_ph_diagnostics(df_ph)          if df_ph          is not None else None
    t_km   = build_extra_km_logrank(df_km_lr)            if df_km_lr       is not None else None
    t_abs  = build_extra_absolute_risks(df_abs_risks)   if df_abs_risks   is not None else None

    t_summary = build_summary(
        t_cont  if t_cont  is not None else pd.DataFrame(),
        t_s2    if t_s2    is not None else pd.DataFrame(),
        t_s4    if t_s4    is not None else pd.DataFrame(),
    )

    # ── Write workbook ─────────────────────────────────────────────────
    out_path = out_dir / "Cox_Formal_Tables.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Supplementary tables (manuscript numbering) ────────────────
        _write_sheet(writer, t_s2,   "ST2 - RBD groups",
                     "Supp Table 2: RBD-only Cox HRs (Model A, percentile_3g, all outcomes)")
        _write_sheet(writer, t_s3,   "ST3 - Threshold stability",
                     "Supp Table 3: RBD threshold stability (PD only, top 5/10/15%)")
        _write_sheet(writer, t_s4,   "ST4 - Prodromal only",
                     "Supp Table 4: Prodromal-only HRs (Model B, PD only, FDR-corrected)")
        _write_sheet(writer, t_s5,   "ST5 - Additive",
                     "Supp Table 5: Additive model (Model C, PD only, percentile_3g)")
        _write_sheet(writer, t_s6,   "ST6 - C-index",
                     "Supp Table 6: Incremental C-index by prodromal marker (PD only)")
        _write_sheet(writer, t_s7,   "ST7 - Interaction",
                     "Supp Table 7: Interaction terms (Model D, PD only, percentile_3g)")
        _write_sheet(writer, t_s9,   "ST9 - Competing risk",
                     "Supp Table 9: Competing-risk cause-specific Cox (Model E, PD only)")
        _write_sheet(writer, t_s10,  "ST10 - Lag sensitivity",
                     "Supp Table 10: Lag sensitivity — events within 2y excluded (PD only)")
        _write_sheet(writer, t_s11,  "ST11 - HES sensitivity",
                     "Supp Table 11: HES-active subcohort sensitivity (PD only, gap ≤ 4y)")
        _write_sheet(writer, t_s12a, "ST12a - RBD spline LRT",
                     "Supp Table 12a: RBD spline LRT (non-linearity test, PD only)")
        _write_sheet(writer, t_s12b, "ST12b - Prodromal spline LRT",
                     "Supp Table 12b: Prodromal spline LRT (linearity test, PD only)")
        # ── Extra / support tables ─────────────────────────────────────
        _write_sheet(writer, t_cont, "Extra - RBD continuous",
                     "RBD continuous HR per 1-SD (all outcomes) — for Fig 1b forest plot")
        _write_sheet(writer, t_cif,  "Extra - CIF vs KM",
                     "CIF vs 1−KM comparison at 5- and 10-year horizons (PD only)")
        _write_sheet(writer, t_abs,  "Extra - Absolute risks",
                     "Absolute cumulative incidence by RBD group (PD only)")
        _write_sheet(writer, t_ph,   "Extra - PH diagnostics",
                     "Proportional hazards Schoenfeld test (PD only, Model A)")
        _write_sheet(writer, t_km,   "Extra - KM log-rank",
                     "KM log-rank summary by prodromal marker (PD only, percentile_3g)")
        _write_sheet(writer, t_summary, "Summary",
                     "Summary: Key effect estimates across models")

    print(f"\nWritten: {out_path}")

    # ── Console preview ────────────────────────────────────────────────
    sep = "=" * 70
    if t_s2 is not None:
        print(f"\n{sep}\n  SUPP TABLE 2 — RBD group HRs (all outcomes, percentile_3g)\n{sep}")
        print(t_s2.to_string(index=False))

    if t_s4 is not None:
        print(f"\n{sep}\n  SUPP TABLE 4 — Prodromal only (PD only, FDR)\n{sep}")
        print(t_s4.to_string(index=False))

    if t_cont is not None:
        print(f"\n{sep}\n  EXTRA — RBD continuous per 1-SD\n{sep}")
        print(t_cont.to_string(index=False))


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(Path(sys.argv[1]))
    else:
        main()
